from openpyxl import load_workbook

path = r'D:\Docs\Works\InProgress\Thesis\Data Analysis\test.xlsx'
wb = load_workbook(path)
ws = wb.active

print(ws['af2'].value)
# Clean the Dates and Time (M, N, O, 0:50-1:40)
'''
Dates depend on the time
'''
# Clean the OSs
'''
Android Linux 10,
Android Linux 8.1.0
iPhone iOS 13.6 -> Unknown Browser
iPhone iOS 14.4 -> Unknown Browser
iPhone iOS 13.6.1 -> Unknown Browser
Windows Wechat -> Chrome 53.0.2785.143
Windows NT 10.0
'''
# Clean the Browsers (P)
'''
Unknown Browser
Chrome 53.0.2785.143

Chrome 78.0.3904.62
Chrome 77.0.3865.120
Chrome 88.0.4324.181
Firefox 85.0
'''

# Step 1. Change IPs and Locations
# Step 2. Change machines and user-agents
# Step 3. Change times to random from 0:50 to 1:40
# Step 4. Count that amount of time from the first date and put it in the second
