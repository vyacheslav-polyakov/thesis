from openpyxl import load_workbook

path = r'D:\Docs\Works\InProgress\Thesis\Data Analysis\test.xlsx'
wb = load_workbook(path)
ws = wb.active

textfile = open('ips.txt', 'r', encoding='utf-16')
left = 501
row = 1
for entry in textfile:
    row += 1
    arr = entry.split(' ')
    ip = arr[0]
    state = arr[1]
    city = arr[3]
    ws['b'+str(row)] = ip
    ws['g'+str(row)] = city
    ws['h'+str(row)] = state
    print('Row {} out of 501'.format(str(row-1)))
    if row == 501:
        break

textfile.close()
wb.save('result.xlsx')




























