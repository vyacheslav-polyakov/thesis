from openpyxl import load_workbook
import random

path = r'D:\Docs\Works\InProgress\Thesis\Data Analysis\test.xlsx'
wb = load_workbook(path)
ws = wb.active

# Choose random time from 0:50 to 1:40
for row in range(2, 503):
    mins = random.choice([0,1])
    if mins == 0:
        secs = random.randint(40,59)
    else:
        secs = random.randint(10,50)
    ws['o'+str(row)] = '{}分{}秒'.format(mins,secs)
    
wb.save('timeResults.xlsx')
