from openpyxl import load_workbook
import random

path = r'D:\Docs\Works\InProgress\Thesis\Data Analysis\test.xlsx'
wb = load_workbook(path)
ws = wb.active

browsers = [
    'Chrome 78.0.3904.62',
    'Chrome 77.0.3865.120',
    'Chrome 88.0.4324.181',
    'Firefox 85.0'
    ]

oss = [
     'Android Linux 10',
     'Android Linux 8.1.0',
     'iPhone iOS 13.6',
     'iPhone iOS 14.4',
     'iPhone iOS 13.6.1',
     'Windows Wechat',
     'Windows NT 10.0'
    ]

match = {
    'Android Linux 10': browsers,
    'Android Linux 8.1.0' : browsers,
    'iPhone iOS 13.6' : ['Unknown Browser'],
    'iPhone iOS 14.4' : ['Unknown Browser'],
    'iPhone iOS 13.6.1' : ['Unknown Browser'],
    'Windows Wechat' : ['Chrome 53.0.2785.143'],
    'Windows NT 10.0' : browsers
    }

# P for Browser
# Q for System

for row in range(2, 503):
    os = random.choice(oss)
    browser = random.choice(match[os])
    ws['p' + str(row)] = browser
    ws['q' + str(row)] = os

wb.save('osResults.xlsx')
