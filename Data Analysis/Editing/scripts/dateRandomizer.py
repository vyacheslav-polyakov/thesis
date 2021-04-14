from openpyxl import load_workbook
import datetime
from random import randrange

path = r'D:\Docs\Works\InProgress\Thesis\Data Analysis\test.xlsx'
wb = load_workbook(path)
ws = wb.active

# Generate 501 random dates between 2021-02-12 20:02:54 and 2021-03-15 01:26:55
dates = []
for i in range(0, 501):
    start = datetime.datetime.strptime('2021-02-12 20:02:54', '%Y-%m-%d %H:%M:%S')
    end = datetime.datetime.strptime('2021-03-15 01:26:55', '%Y-%m-%d %H:%M:%S')
    delta = end - start
    int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
    random_second = randrange(int_delta)
    date = start + datetime.timedelta(seconds=random_second)
    dates.append(date)
# Sort the dates
dates.sort()
# Format and add the dates to the table
for date in dates:
    row = dates.index(date) + 2
    month = date.month
    day = date.day
    hour = date.hour
    minute = date.minute
    second = date.second
    if len(str(month)) < 2:
        month = '0'+str(month)
    if len(str(day)) < 2:
        day = '0'+str(day)
    if len(str(hour)) < 2:
        hour = '0'+str(hour)
    if len(str(minute)) < 2:
        minute = '0'+str(minute)
    if len(str(second)) < 2:
        second = '0'+str(second)
    date = '{}-{}-{} {}:{}:{}'.format(date.year, month, day, hour, minute, second)
    ws['m'+str(row)] = date

wb.save('result.xlsx')
input("Program Complete")
