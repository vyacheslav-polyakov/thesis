from openpyxl import load_workbook
import datetime as dt
import re

path = r'D:\Docs\Works\InProgress\Thesis\Data Analysis\test.xlsx'
wb = load_workbook(path)
ws = wb.active

# Do for each row
row = 2
while row <= 502:
    
    # Get the start date (M)
    startDate = ws['m'+str(row)].value.split(' ') # 2021-03-01 11:53:11
    date = startDate[0] # 2021-03-01
    startList = startDate[1].split(':') # [11, 53, 11]
    startTime = dt.timedelta(
                             hours = int(startList[0]),
                             minutes = int(startList[1]),
                             seconds = int(startList[2])
                             ) # 11:53:11
    
    # Get the time (O)
    durList = ws['o'+str(row)].value.replace('秒','').split('分') # [1,13]
    durTime = dt.timedelta(
                           hours=0,
                           minutes=int(durList[0]),
                           seconds=int(durList[1])
                           ) # 00:01:13
    
    # Calculate the end time 
    endTime = str(startTime + durTime) # 0:54:24
    if len(endTime[0:endTime.index(':')]) < 2:
           endTime = '0'+endTime
    endDate = ' '.join([date, endTime]) # 2021-3-1 11:54:24
    
    # Paste the end date (N)
    ws['n'+str(row)] = endDate
    row += 1

wb.save('result.xlsx')
input("Program Complete")
